# -*- coding: utf-8 -*-
"""
반송 시나리오 6종 - 동작 단위 정의서 생성
  PlcAddressMap.xml(단일 기준)에서 단계·주소를 그대로 파생한다. 주소를 손으로 적지 않는다.
    · 구 ECS 표기 : 문서(PPT/구 ECS TB_OBSERVABLE) 표기.  M워드.비트 / D워드 / R워드(16진 해석)
    · 실제 주소   : XGT 실전송 주소.  %MX(비트) / %DB(=워드x2) / %RB(=워드x2)
"""
import sys, io, xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XML = 'TASK/WCS_TASK_CV_BIN/7_DeviceMap/PlcAddressMap.xml'
ROOT = ET.parse(XML).getroot()
R_HEX = (ROOT.get('rAddrMode', 'HEX').upper() == 'HEX')

GRP = {g.get('type'): g for g in ROOT.findall('EquipGroup')}
EQUIP = {}
for e in ROOT.iter('Equip'):
    EQUIP[(e.get('type'), e.get('no'))] = e.attrib


def blk(gtype, bname):
    for b in GRP[gtype].findall('Block'):
        if b.get('name') == bname:
            return b
    return None


def track_of(gtype, no, slot):
    a = EQUIP.get((gtype, no))
    if not a or slot is None:
        return ''
    if a.get('trackOrder'):
        order = a['trackOrder'].split(',')
        return order[slot] if slot < len(order) else ''
    fr = int(a['frTrack'])
    return str(fr + slot)


CRANE_NO = 5          # 사용자 제시 표 기준 : 22번 트랙 → S/C #5 입고
CRANE = {c.get('no'): c.attrib for c in ROOT.iter('Crane')}


def resolve_no(expr):
    """설비번호 식 치환 : $ = 호기, $inCv / $outCv = 그 호기의 통로 C/V (XML CraneMap 기준)."""
    if expr is None:
        return None
    e = str(expr).strip()
    c = CRANE.get(str(CRANE_NO), {})
    if e == '$':
        return str(CRANE_NO)
    if e.lower() == '$incv':
        return c.get('inCv')
    if e.lower() == '$outcv':
        return c.get('outCv')
    return e


def addr_of(st):
    """(구 ECS 표기, 실제 주소, 디바이스) 반환. 주소가 없는 단계는 ('','','')."""
    gtype, no = st.get('equip'), resolve_no(st.get('no'))
    bname, sname = st.get('block'), st.get('signal')
    if not (gtype and no and bname and sname):
        return '', '', ''
    b = blk(gtype, bname)
    if b is None:
        return '', '', ''
    dev = b.get('device')
    base = int(b.get('origin')) + (int(no) - int(GRP[gtype].get('numberFrom'))) * int(b.get('stride'))
    slot = int(st.get('slot')) if st.get('slot') is not None else None

    off = None
    per = 1
    for s in list(b):
        if s.get('name') == sname:
            off = int(s.get('offset'))
            if s.tag == 'SignalArray':
                per = int(s.get('perSlot', 1))
                off += (slot or 0) * per
            break
    if off is None:
        return '', '', ''

    if dev == 'M':
        a = base + off
        return 'M%03d.%d' % (a // 16, a % 16), '%%MX%d' % a, 'M'
    if dev == 'D':
        a = base + off
        return 'D%04d' % a, '%%DB%d' % (a * 2), 'D'
    # R : 문서표기(10진 자릿수)를 16진으로 해석
    doc = base + off + (slot or 0) * int(b.get('perSlotWords', 2))
    real = int(str(doc), 16) if R_HEX else doc
    return 'R%04d' % doc, '%%RB%d' % (real * 2), 'R'


KIND_KOR = {'observe': '설비 → WCS (관측)',
            'force':   'WCS → 설비 (지시/ACK)',
            'word':    '데이터 영역 (D/R)'}

# src 가 비어 있는 단계의 처리 주체 추정 규칙
DEFAULT_SRC = {
    'observe': 'EQP_SIM(설비/PLC) 가 발생 · WCS_TASK_CV CvThread.cs::CvStatusScenario 가 판독',
    'word':    'WCS_TASK_CV CvThread.cs::CvStatusScenario (판독) / CvTrackingWrite (기록)',
    'force':   'WCS_TASK_CV CvThread.cs::CvEventCheck',
}



# ── 작업상태 / HOST 전문 계층 (PLC 신호 계층 아래에 덧붙인다) ───────────────
FLOW_IN = [
 ('평상시', '-', '-', 'S 상태보고(30초 주기 + 상태변경시)',
  'WCS_TASK_HOST CCliWork.cs::GetStatusReport'),
 ('입고대 신호 ON', '-', '-', 'S 상태보고 (101 입고대 ON)',
  'EQP_SIM ConveyorSim.cs::UpdateStationSignals → CvThread.cs::CvStatusScenario → CV_DATA.STO_READY_RD'),
 ('HOST 입고 작업 수신', '신규(99)', '작업번호 채번', 'O 작업지시 수신 → o ACK',
  'WCS_TASK_HOST CSrvWork.cs (O 전문 파싱) → JOB_MST INSERT(99)'),
 ('작업 초기화', 'CV 구동대기(10)', '-', '-',
  'IO_TASK cThread_SCH.cs::AcceptNewJob  (입고 99→10 / 출고 99→20)'),
 ('입고 작업 CV 에 기록', 'CV 구동지시(11) → CV 구동중(15)', '트래킹 R영역 기록', '-',
  'IO_TASK cThread_SCH.cs::DriveCV / RunCV  ·  WCS_TASK_CV CvThread.cs::CvTrackingWrite'),
 ('RGV 출발/도착 H/S 판정', '-', '-', '-',
  'WCS_TASK_CV CvThread.cs::CvStatusScenario (bPickOn/bDropOn) → CV_DATA.RTV_DEPARTHS/ARRIVEHS_READY_RD'),
 ('RGV 작업지시', 'RGV 구동지시(31) → 구동중(35)', '-', '-',
  'IO_TASK cThread_SCH.cs::DriveRGV / RunRGV  ·  WCS_TASK_CV VehThread.cs::ConsumeCommands'),
 ('RGV 작업 완료', 'SC 구동대기(20)로 인계', '-', '-',
  'IO_TASK cThread_SCH.cs::CompleteRGVReal  (완료 직후 같은 주기에 DriveSC 재호출 - 대기 미체류)'),
 ('입고 H/S 신호 ON', '-', '-', '-',
  'WCS_TASK_CV CvThread.cs::CvStatusScenario → CV_DATA.STOHS_READY_RD'),
 ('SC 작업지시', 'SC 구동지시(21) → 구동중(25)', '-', '-',
  'IO_TASK cThread_SCH.cs::DriveSC / RunSC  ·  WCS_TASK_CV VehThread.cs::ConsumeCommands'),
 ('SC 작업 완료', 'SC 구동완료 → 작업완료(29)', '-', 'F 입고작업 완료보고',
  'IO_TASK cThread_SCH.cs::CompleteSC  →  WCS_TASK_HOST CCliWork.cs::GetJobCompleteReport'),
 ('HOST 완료 응답 수신', '작업 삭제', '-', 'f ACK 수신',
  'WCS_TASK_HOST CCliWork.cs (RequestSrv 응답 대기) → frmMain.cs::DeleteJobMst  → JOB_MST 삭제 / JOB_MST_HIS 이관'),
]

FLOW_OUT = [
 ('평상시', '-', '-', 'S 상태보고(30초 주기 + 상태변경시)',
  'WCS_TASK_HOST CCliWork.cs::GetStatusReport'),
 ('HOST 출고 작업 수신', '신규(99)', '작업번호 채번', 'O 작업지시 수신 → o ACK',
  'WCS_TASK_HOST CSrvWork.cs (O 전문 파싱) → JOB_MST INSERT(99)'),
 ('작업 초기화', 'SC 구동대기(20)', '-', '-',
  'IO_TASK cThread_SCH.cs::AcceptNewJob  (출고 99→20)'),
 ('SC 작업지시', 'SC 구동지시(21) → 구동중(25)', '-', '-',
  'IO_TASK cThread_SCH.cs::DriveSC / RunSC  ·  WCS_TASK_CV VehThread.cs::ConsumeCommands'),
 ('SC 하역 완료', 'RGV 구동대기(30)로 인계', '-', '-',
  'IO_TASK cThread_SCH.cs::CompleteSC  (완료 직후 같은 주기에 DriveRGV 재호출 - 대기 미체류)'),
 ('RGV 작업지시/완료', 'RGV 31 → 35 → CV 구동대기(10)', '-', '-',
  'IO_TASK cThread_SCH.cs::DriveRGV / RunRGV / CompleteRGVReal'),
 ('겸용대 방향 정합', '-', '-', 'M 모드변경(상위 지시 시)',
  'IO_TASK cThread_SCH.cs::SyncDualCvDirection → RequestCvDirection  ·  WCS_TASK_CV CvThread.cs (CMD_RQ_ID=DIR, IsDualCvDirChangeHeld)'),
 ('CV 반출 지시', 'CV 구동지시(11) → 구동중(15)', '-', '-',
  'IO_TASK cThread_SCH.cs::DriveCV / RunCV'),
 ('출고대 신호 ON', '출고 H/S 도착(22)', '-', 'F 도착보고 (StepCount=1)',
  'EQP_SIM ConveyorSim.cs::UpdateStationSignals (화물O + 데이터O) → CvThread → CV_DATA.RET_READY_RD → IO_TASK cThread_SCH.cs::ReportOutStationArrival'),
 ('HOST 도착 응답 수신', '작업 삭제', '-', 'f ACK 수신',
  'WCS_TASK_HOST CCliWork.cs::GetLoadArrivalReport (RequestSrv 응답 대기) → frmMain.cs::DeleteJobMst'),
 ('지게차 반출', '-', '트래킹 제거', '-',
  'EQP_SIM ConveyorSim.cs : 신호 ON +OUT_REMOVE_MS 화물감지 OFF → +OUT_TRACK_CLEAR_MS 트래킹 삭제'),
]

FLOW_HDR = ['단계', 'WCS 작업상태', 'WCS 데이터', 'HOST 전문', '처리 시스템 / 파일 :: 함수']


def write_flow(ws, row, is_in):
    ws.cell(row, 1, '■ 작업상태 · HOST 전문 계층 (PLC 신호 계층과 짝을 이루는 상위 흐름)').font = Font(bold=True, size=11)
    row += 1
    for c, h in enumerate(FLOW_HDR, 1):
        cell = ws.cell(row, c, h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor='BDD7EE')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    row += 1
    for f in (FLOW_IN if is_in else FLOW_OUT):
        for c, v in enumerate(f, 1):
            cell = ws.cell(row, c, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center', wrap_text=(c in (1, 5)))
        row += 1
    return row

HDR = ['번호', '구분', '동작', '설비', 'Track',
       'PLC 영역\n구 ECS 표기', 'PLC 영역\n실제 주소(XGT)',
       '처리 시스템 / 파일 :: 함수']

thin = Side(style='thin', color='999999')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    wb = Workbook()
    wb.remove(wb.active)
    scs = ROOT.find('Scenarios').findall('Scenario')
    for sc in scs:
        title = sc.get('title')
        name = ('S' + sc.get('id')).replace('.', '_')[:31]
        ws = wb.create_sheet(name)

        ws.cell(1, 1, title).font = Font(bold=True, size=13)
        ws.cell(2, 1, '기준 : PlcAddressMap.xml v%s (rAddrMode=%s) · S/C #%d 기준으로 설비번호 식($ / $inCv / $outCv)을 치환했다. 주소는 손으로 적지 않고 이 파일에서 계산했다.'
                % (ROOT.get('version'), ROOT.get('rAddrMode'), CRANE_NO))
        ws.cell(2, 1).font = Font(size=9, color='666666')
        ws.cell(3, 1, 'M 비트 : 워드=비트/16 · D 워드 : %DB=워드x2 · R 트래킹 : 문서표기를 16진 해석 후 %RB=워드x2')
        ws.cell(3, 1).font = Font(size=9, color='666666')

        hr = 5
        for c, h in enumerate(HDR, 1):
            cell = ws.cell(hr, c, h)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill('solid', fgColor='D9D9D9')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER

        r = hr + 1
        for i, st in enumerate(sc.findall('Step'), 1):
            kind = st.get('kind', '')
            legacy, real, dev = addr_of(st)
            slot = int(st.get('slot')) if st.get('slot') is not None else None
            rno = resolve_no(st.get('no'))
            eq = '%s#%s' % (st.get('equip', ''), rno or '') if st.get('equip') else ''
            src = st.get('src') or DEFAULT_SRC.get(kind, '')
            vals = [i, KIND_KOR.get(kind, kind), st.get('desc', ''), eq,
                    track_of(st.get('equip'), rno, slot),
                    legacy, real, src]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=(c in (3, 8)))
                cell.font = Font(size=10)
                if kind == 'force':
                    cell.fill = PatternFill('solid', fgColor='FFF2CC')   # WCS 가 내는 것
                elif kind == 'word':
                    cell.fill = PatternFill('solid', fgColor='E2EFDA')   # 데이터 영역
            r += 1

        is_in = ('입고' in title)
        write_flow(ws, r + 2, is_in)

        for col, w in zip('ABCDEFGH', (6, 22, 46, 9, 8, 15, 14, 52)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A6'
    out = '반송시나리오6종_동작단위_영역및처리_20260830.xlsx'
    wb.save(out)
    print('saved:', out, '| sheets:', len(scs))


build()
