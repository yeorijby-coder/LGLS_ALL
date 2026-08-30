# -*- coding: utf-8 -*-
"""V1.5 시나리오 영역 xlsx  vs  현재 시스템(PlcAddressMap.xml) 대조."""
import sys, io, re, xml.etree.ElementTree as ET
from openpyxl import load_workbook
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = '시나리오 영역_시나리오그림포함_V1.5_20260828.xlsx'
ROOT = ET.parse('TASK/WCS_TASK_CV_BIN/7_DeviceMap/PlcAddressMap.xml').getroot()
R_HEX = (ROOT.get('rAddrMode', 'HEX').upper() == 'HEX')
GRP = {g.get('type'): g for g in ROOT.findall('EquipGroup')}
CRANE = {c.get('no'): c.attrib for c in ROOT.iter('Crane')}
CN = '1'   # 시트가 SC #1 기준

def blk(t, n):
    for b in GRP[t].findall('Block'):
        if b.get('name') == n: return b
    return None

def rno(e):
    if e is None: return None
    s = str(e).strip(); c = CRANE.get(CN, {})
    if s == '$': return CN
    if s.lower() == '$incv': return c.get('inCv')
    if s.lower() == '$outcv': return c.get('outCv')
    return s

def addr(st):
    t, n = st.get('equip'), rno(st.get('no'))
    bn, sn = st.get('block'), st.get('signal')
    if not (t and n and bn and sn): return '', ''
    b = blk(t, bn)
    if b is None: return '', ''
    dev = b.get('device')
    base = int(b.get('origin')) + (int(n) - int(GRP[t].get('numberFrom'))) * int(b.get('stride'))
    slot = int(st.get('slot')) if st.get('slot') is not None else None
    off, per = None, 1
    for s in list(b):
        if s.get('name') == sn:
            off = int(s.get('offset'))
            if s.tag == 'SignalArray':
                per = int(s.get('perSlot', 1)); off += (slot or 0) * per
            break
    if off is None: return '', ''
    leg = b.get('legacy', '')
    lstride = int(b.get('legacyStride', b.get('stride', 1)))
    if dev == 'M':
        a = base + off
        return 'M%03d.%d' % (a // 16, a % 16), '%%MX%d' % a
    if dev == 'D':
        a = base + off
        lb = int(re.sub(r'\D', '', leg) or 0) + (int(n) - int(GRP[t].get('numberFrom'))) * lstride + off
        return 'D%04d' % lb, '%%DW%d' % a
    doc = base + off + (slot or 0) * int(b.get('perSlotWords', 2))
    real = int(str(doc), 16) if R_HEX else doc
    return 'R%04d' % doc, '%%RB%d' % (real * 2)

SUB = {'observe': 'EQP BIT', 'force': 'WCS BIT ON', 'word': 'WCS 값'}

def xml_steps(sid):
    for sc in ROOT.find('Scenarios').findall('Scenario'):
        if sc.get('id') == sid:
            out = []
            for st in sc.findall('Step'):
                lg, rl = addr(st)
                out.append((st.get('desc', ''), lg, rl, st.get('kind'), st.get('src') or ''))
            return out
    return []

SHEETS = [('입고 TR #22 => SC #1', '1'), ('출고 SC #1 => TR #22', '2'),
          ('입고 TR #24 => SC #1', '3-1'), ('출고 SC #1 => TR #26', '3-2'),
          ('입고 TR #30 => SC #1', '4-1'), ('출고 SC #1 => TR #29', '4-2')]

wb = load_workbook(SRC)
total_diff = 0
for sh, sid in SHEETS:
    ws = wb[sh]
    xls = []
    for r in range(4, ws.max_row + 1):
        num = ws.cell(r, 10).value
        if num is None: continue
        desc = (ws.cell(r, 13).value or '').strip()
        a14 = (ws.cell(r, 14).value or '')
        m = re.search(r'([MDR]\d{3,4}(?:\.\d)?)\s+(%\w+\d+)', str(a14))
        xls.append((desc, m.group(1) if m else '', m.group(2) if m else '',
                    (ws.cell(r, 19).value or ''), (ws.cell(r, 20).value or ''), (ws.cell(r, 21).value or '')))
    xs = xml_steps(sid)
    print('=== %-22s  xlsx=%-3d  system=%-3d' % (sh, len(xls), len(xs)))
    n = max(len(xls), len(xs))
    diff = 0
    for i in range(n):
        a = xls[i] if i < len(xls) else ('(없음)', '', '', '', '', '')
        b = xs[i] if i < len(xs) else ('(없음)', '', '', '', '')
        if a[1] != b[1] or a[2] != b[2] or a[0][:18] != b[0][:18]:
            diff += 1
            if diff <= 4:
                print('   #%-3d xlsx[%s | %s %s]' % (i + 1, a[0][:34], a[1], a[2]))
                print('        sys [%s | %s %s]' % (b[0][:34], b[1], b[2]))
    if diff == 0: print('   → 영역·순서·설명 일치')
    else: print('   → 불일치 %d건' % diff)
    total_diff += diff
print()
print('총 불일치 :', total_diff)
